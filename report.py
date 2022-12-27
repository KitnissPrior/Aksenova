import openpyxl
from openpyxl.styles import numbers, Font
from openpyxl.styles.borders import Border, Side
import matplotlib.pyplot as plt
import numpy as np
import pdfkit
from jinja2 import Environment, FileSystemLoader
import statistics as stats

class Report:
    """
    Класс для представления отчета

    Attributes:
        job (str): Выбранная профессия
        wb (WorkBook): содержимое excel-файла с отчетом
        years_table_headers (list): заголовки таблицы со статистикой по годам
        years_table (list): таблица со статистикой по годам
        cities_table_headers (list): заголовки таблицы со статистикой по городам
        cities_salary_table (list): таблица со статистикой зарплат по городом
        cities_vacancy_table (list): таблица с данными по долям вакансий по городом
    """
    def __init__(self, job):
        """Инициализирует объект Report

        Args:
            job (str): Выбранная профессия
        """
        self.job = job
        self.wb = openpyxl.Workbook()
        self.years_table_headers = []
        self.years_table = []
        self.cities_table_headers = []
        self.cities_salary_table = []
        self.cities_vacancy_table = []

    def set_bold_titles(self):
        """Устанавливает жирный шрифт заголовков таблиц на всех листах excel-файла"""
        for name in self.wb.sheetnames:
            sheet = self.wb[name]
            for i in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=1, column=i)
                cell.font = Font(bold=True)

    def set_thin_borders(self):
        """Устанавливает тонкие границы для таблиц на всех листах excel-файла"""
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        for name in self.wb.sheetnames:
            sheet = self.wb[name]
            for row in sheet.iter_rows():
                for cell in row:
                    if ((name == 'Статистика по городам' and cell.column != 3)
                            or name == 'Статистика по годам'):
                        cell.border = thin_border

    def add_years_statistics(self, years_statistics, titles, sheet_num):
        """Добавляет в рабочую книгу excel статистику по годам

        Args:
            years_statistics (dict): данные статистики по годам
            titles (list): заголовки для будущей таблицы со статистикой
            sheet_num (int): номер текущего листа в excel-книге
        """
        sheet = self.wb.worksheets[sheet_num]
        sheet.append((titles))
        self.years_table_headers = titles

        years = list(years_statistics.values())
        values_count = len(years[0])
        for i in range(values_count):
            values = []
            values.append(list(years[0].keys())[i])
            for j in range(len(years)):
                values.append(list(years[j].values())[i])
            sheet.append(values)
            self.years_table.append(values)

    def set_table_format(self):
        """Настраивает формат таблиц в excel-файле"""
        self.set_bold_titles()
        self.set_thin_borders()
        self.set_columns_width()

    def set_percentage_format(self, sheet_number, rows_count):
        """Устанавливает процентный формат данных для столбца с долей вакансий по городам
        
        Args:
            sheet_number (int): номер текущего листа в excel-файле
            rows_count (int): количество строк в столбце с долей вакансий по городам
        """
        sheet = self.wb.worksheets[sheet_number]
        for i in range(1, rows_count):
            cell = sheet.cell(row=i, column=4)
            cell.number_format = numbers.FORMAT_PERCENTAGE_00

    def add_cities_statistics(self, cities_statistics, titles, sheet_num):
        """Добавляет в рабочую книгу excel статистику по городам

        Args:
            cities_statistics (dict): данные статистики по годам
            titles (list): заголовки для будущей таблицы со статистикой
            sheet_num (int): номер текущего листа в excel-книге
        """
        sheet = self.wb.worksheets[sheet_num]
        sheet.append((titles))
        self.cities_table_headers = titles

        cities = list(cities_statistics.values())
        values_count = len(cities[0])
        for i in range(values_count):
            values = []
            for j in range(len(cities)):
                items = list(cities[j].items())[i]
                values.append(items[0])
                values.append(items[1])
            sheet.append(values)
            self.cities_salary_table.append(values[0:2])
            self.cities_vacancy_table.append(values[2::])

        self.set_percentage_format(1, values_count + 2)
        sheet.insert_cols(3, 1)

    def set_columns_width(self):
        """Настраивает ширину столбцов для всех листов excel-файла,
           чтобы каждый столбец вмещал самую длинную строку в столбце
        """
        for name in self.wb.sheetnames:
            sheet = self.wb[name]
            dims = {}
            for row in sheet.rows:
                for cell in row:
                    length = len(str(cell.value)) + 2
                    width = max((dims.get(cell.column_letter, 0), length))
                    dims[cell.column_letter] = width
            for col, value in dims.items():
                sheet.column_dimensions[col].width = value

    def generate_excel(self, years, cities):
        """Генерирует excel-файл с отчетом по статистике

        Args:
            years (dict): статистика вакансий по годам
            cities (dict): статистика вакансий по городам
        """
        sheet_1 = self.wb.active
        sheet_1.title = 'Статистика по годам'
        sheet_2 = self.wb.create_sheet('Статистика по городам')

        years_titles = ['Год', 'Средняя зарплата', f"Средняя зарплата - {self.job}",
                        'Количество вакансий', f"Количество вакансий - {self.job}"]
        cities_titles = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']

        self.add_years_statistics(years, years_titles, 0)
        self.add_cities_statistics(cities, cities_titles, 1)

        self.set_table_format()
        self.wb.save('report.xlsx')

    def plot_bar_chart(self, ax, values1, values2, title, xlabels, label1, label2):
        """Строит гистограмму статистики для двух графиков

        Args:
            ax (matplotlib.axes.Axes): область изображения, которую будет занимать гистограмма
            values1 (list): значения первого графика по оси Ох
            values2 (list): значения второго графика по оси Ох
            title (str): заголовок гистограммы
            xlabels (list): подписи по оси Ох
            label1 (str): легенда для значений первого графика
            label2 (str): легенда для значений второго графика
        """
        x = np.arange(len(xlabels))
        width = 0.35

        ax.bar(x - width / 2, values1, width, label=label1)
        ax.bar(x + width / 2, values2, width, label=label2)

        ax.set_title(title)
        ax.set_xticks(x, xlabels, rotation=90)

        ax.legend(loc="upper left")
        ax.grid(True, axis="y")
        ax.tick_params(axis='both', labelsize=8)

    def plot_salaries_by_years_chart(self, ax, years_statistics):
        """Строит график уровня зарплат по годам

           Args:
               ax (matplotlib.axes.Axes): область изображения, которую будет занимать гистограмма
               years_statistics (dict): данные статистики по годам
        """
        years = list(years_statistics['salary_all'].keys())
        salaries_all = list(years_statistics['salary_all'].values())
        salaries_job = list(years_statistics['salary_job'].values())

        self.plot_bar_chart(ax, salaries_all, salaries_job,
                            'Уровень зарплат по годам', years,
                            'средняя з/п', f"з/п {self.job.lower()}")

    def plot_vacancies_by_years_chart(self, ax, years_statistics):
        """Строит график количества вакансий по годам

           Args:
                ax (matplotlib.axes.Axes): область изображения, которую будет занимать гистограмма
                years_statistics (dict): данные статистики по годам
        """
        years = list(years_statistics['number_all'].keys())
        number_all = list(years_statistics['number_all'].values())
        number_job = list(years_statistics['number_job'].values())

        self.plot_bar_chart(ax, number_all, number_job,
                            'Количество вакансий по годам', years,
                            'Количество вакансий',
                            f"Количество вакансий\n{self.job.lower()}")

    def format_cities_labels(self, cities):
        """Форматирует подписи с названиями городов, добавляя переносы слов
        для городов, состоящих из нескольких слов, разделенных пробелом или дефисом

           Args:
               cities (list): названия городов
        """
        result = []
        for city in cities:
            if '-' in city:
                parts = city.split('-')
                city = f"{parts[0]}-\n{parts[1]}"
                if len(parts) > 2:
                    city = f"{city}-\n{parts[2]}"
            if ' ' in city:
                city = '\n'.join(city.split(' '))
            result.append(city)
        return result

    def plot_salaries_by_cities_chart(self, ax, cities_statistics):
        """Строит горизонтальную столбчатую диаграмму по данным
        об уровне зарплат по городам

            Args:
                ax (matplotlib.axes.Axes): область изображения, которую будет занимать диаграмма
                cities_statistics (dict): данные статистики по городам
        """
        cities = self.format_cities_labels(
            list(cities_statistics['salary'].keys()))
        salaries = list(cities_statistics['salary'].values())
        y = np.arange(len(cities))

        ax.barh(y, salaries, align='center')
        ax.set_yticks(y, labels=cities)

        ax.tick_params(axis='y', labelsize=6)
        ax.invert_yaxis()
        ax.set_title('Уровень зарплат по городам')
        ax.grid(True, axis="x")

    def plot_vacancies_by_cities_chart(self, ax, cities_statistics):
        """Строит круговую диаграмму соотношения количества вакансий по городам

           Args:
               ax (matplotlib.axes.Axes): область изображения, которую будет занимать диаграмма
               cities_statistics (dict): данные статистики по городам
        """
        cities = list(cities_statistics['proportion'].keys())
        vacancies = list(cities_statistics['proportion'].values())

        vacancies.insert(0, 1 - sum(vacancies))
        cities.insert(0, 'Другие')

        ax.tick_params(axis='x', labelsize=6)
        ax.pie(vacancies, labels=cities, textprops={'size': 6})

        ax.set_title('Доля вакансий по городам')

    def generate_image(self, years, cities):
        """Генерирует изображение с графиками статистики

           Args:
               years (dict): статистика по годам
               cities( dict): статистика по городам
        """
        fig, axes = plt.subplots(2, 2)

        self.plot_salaries_by_years_chart(axes[0, 0], years)
        self.plot_vacancies_by_years_chart(axes[0, 1], years)
        self.plot_salaries_by_cities_chart(axes[1, 0], cities)
        self.plot_vacancies_by_cities_chart(axes[1, 1], cities)

        plt.rcParams['font.size'] = 8
        fig.tight_layout()

        plt.savefig('graph.png',dpi=300)

    def fill_pdf_with_data(self, template, config):
        """Заполняет pdf-файл данными по статистике

           Args:
               template (Template): шаблон отчета в формате html
               config (Configuration): настройки конфигурации для преобразования файла из html в pdf
        """
        img_file = 'D:\\ИРИТ\\2 курс\\питон\\Aksenova\\graph.png'

        sheet_names = self.wb.sheetnames
        pdf_template = template.render({'job': self.job, 'img_file': img_file,
                                        'years_title': sheet_names[0],
                                        'years_headers': self.years_table_headers,
                                        'years_table': self.years_table, 'cities_title': sheet_names[1],
                                        'cities_salary_headers': self.cities_table_headers[0:2],
                                        'cities_salary_table': self.cities_salary_table,
                                        'cities_vacancy_headers': self.cities_table_headers[2::],
                                        'cities_vacancy_table': self.cities_vacancy_table})

        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})

    def generate_pdf(self, years, cities):
        """Генерирует отчет в виде pdf-файла с данными статистики по выбранной профессии

           Args:
               years (dict): статистика по годам
               cities( dict): статистика по городам
        """
        self.generate_excel(years, cities)
        self.generate_image(years, cities)

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")
        config = pdfkit.configuration(wkhtmltopdf=r'D:\wkhtmltopdf\bin\wkhtmltopdf.exe')

        self.fill_pdf_with_data(template, config)

def get_report():
    """Создает отчет с данными статистики по выбранной профессии"""
    folder_name = input('Введите название папки: ')
    job = input('Введите название профессии: ')

    data_set = stats.DataSet(folder_name, job)
    data_set.parse_csv()
    years, cities = data_set.connector.get_statistics()

    report = Report(job)
    report.generate_pdf(years, cities)

